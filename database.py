#Library to store and manage device data
import os
import pandas as pd
from openpyxl import load_workbook

data_store = {}  # A dictionary to store device data

def device_id_entry(device_id):
    """
    Initializes the dictionary structure for a new MAC address.
    Args:
        device_id_entry (str): The MAC address of the device.
    """
    if device_id.lower() not in data_store:
        data_store[device_id.lower()] = {
            'Packet Count': 0,
            'Company Name': None,
            'Device ID': None,
            'Button Release Time': None,
            'First Packet Time': None,
            'Time Gap': None,
            'Average RSSI': None,
            'Average Frequency': None,
            'Payload Version': None,
            'FW Version': None,
            'Device Type': None,
            'Time Reference': None,
            'Temperature': None,
            'Event Ordinal Number': None,
            'Digital Signature Matched': None,
            'Data Integrity Check Success Count': None,
            'Error Numbers': None,
            'Error Message': None
        }

def update_attribute(device_id, attribute, value):
    """
    Updates the value of a specific attribute for a given MAC address.
    Args:
        mac (str): The MAC address of the device.
        attribute (str): The attribute to be updated.
        value: The new value for the attribute.
    """
    if device_id.lower() in data_store and attribute in data_store[device_id.lower()]:
        data_store[device_id.lower()][attribute] = value

def get_data(device_id):
    """
    Retrieves the data for a given MAC address.
    Args:
        mac (str): The MAC address of the device.
    Returns:
        dict: The data for the given MAC address.
    """
    return data_store[device_id.lower()]

def save_data_to_excel(device_id, update_notification_callback):
    """
    Saves the data for a given MAC address to an Excel file.
    Args:
        device_id (str): The MAC address of the device.
        update_notification_callback (function): A callback function for notifications.
    """
    # Check if the MAC address exists in the data_store
    if device_id.lower() not in data_store:
        update_notification_callback("No data found for the given Device ID.")
        return
    
    # Retrieve the data for the MAC address
    device_id_data = get_data(device_id.lower())
    
    # Convert the data into a DataFrame
    df = pd.DataFrame([device_id_data])
    
    # Define the Excel file name
    excel_file_name = f"saved_data.xlsx"
    
    # Check if the file exists
    if os.path.exists(excel_file_name):
        # Load the existing workbook
        book = load_workbook(excel_file_name)
        
        # Check if the sheet exists
        if 'Device ID' in book.sheetnames:
            # Use ExcelWriter to write to the existing workbook
            with pd.ExcelWriter(excel_file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Get the last row number in the existing sheet
                startrow = book['Device ID'].max_row
                
                # Convert the dataframe to an Excel object
                df.to_excel(writer, sheet_name='Device ID', startrow=startrow, header=not bool(startrow))
        else:
            # If sheet does not exist, write with header
            with pd.ExcelWriter(excel_file_name, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name='Device ID')
    else:
        # Save the DataFrame to a new Excel file
        with pd.ExcelWriter(excel_file_name, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Device ID')
    
    update_notification_callback(f"testing Data for device id {device_id.lower()}  saved to {excel_file_name} successfully.")
